import json
from pathlib import Path
from openpyxl import load_workbook

src = Path("家計.xlsm")
out = Path("data/kakei.json")

out.parent.mkdir(parents=True, exist_ok=True)

wb = load_workbook(src, data_only=True, read_only=True)

data = {}

for ws in wb.worksheets:
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    data[ws.title] = rows

out.write_text(
    json.dumps(data, ensure_ascii=False, default=str, indent=2),
    encoding="utf-8"
)

print(f"生成完了: {out}")
