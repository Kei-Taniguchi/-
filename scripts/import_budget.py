from pathlib import Path
import json
import os
import re
from datetime import datetime
import openpyxl

ROOT=Path(__file__).resolve().parents[1]
DEFAULT_SOURCE=Path(r"C:\Users\syrup\OneDrive\デスクトップ\家計.xlsm")
BOOK_XLSM=ROOT/"家計.xlsm"
BOOK_XLSX=ROOT/"家計.xlsx"
OUT=ROOT/"data"/"budget-data.json"
MONTH_RE=re.compile(r"^(\d{4})年(\d{1,2})月分$")
RATIO_RE=re.compile(r"(\d+)\s*/\s*(\d+)")
END_RE=re.compile(r"[～~]\s*(\d{4})\s*/\s*(\d+)")
CARD_ALIASES={"イオンカード":"イオンカード","イオン":"イオンカード","セゾンカード":"セゾンカード","セゾン":"セゾンカード","PayPayカード":"PayPayカード","PayPay":"PayPayカード","UFJ Nicos":"UFJ Nicos","Nicos":"UFJ Nicos","ニコス":"UFJ Nicos","楽天カード":"楽天カード","楽天":"楽天カード"}
OFFICIAL_SCHEDULES=[{"name":"イオンカード","引落":"翌月2日","締め日":"毎月10日","source":"https://www.aeon.co.jp/service/shopping/"},{"name":"セゾンカード","引落":"翌月4日","締め日":"毎月10日","source":"https://www.saisoncard.co.jp/creditcard/"},{"name":"PayPayカード","引落":"翌月27日","締め日":"毎月末日","source":"https://www.paypay-card.co.jp/service/000173.html"},{"name":"UFJ Nicos","引落":"当月27日","締め日":"毎月5日","source":"https://www.cr.mufg.jp/carduse/before/index.html"},{"name":"楽天カード","引落":"翌月27日","締め日":"毎月末日","source":"https://www.rakuten-card.co.jp/support/guide/"}]
OFFICIAL_SERVICES=[{"name":"モビット","引落":"毎月5日・15日・25日・末日のいずれか（契約時に選択）","締め日":"カードの締め日は設定なし（返済期日方式）","source":"https://www.smbc-card.com/nyukai/loan/conditions.jsp"},{"name":"千葉銀","引落":"毎月1日（銀行休業日は翌営業日）","締め日":"カードの締め日は設定なし（カードローン返済方式）","source":"https://www.chibabank.co.jp/s/lp/card_loan/ad/lp_01/"}]

def number(v):return isinstance(v,(int,float)) and not isinstance(v,bool)
def parse_ratio(*values):
    for v in values:
        if v is not None:
            m=RATIO_RE.search(str(v))
            if m:return int(m.group(1)),int(m.group(2))
    return None
def is_completed_gray_row(ws,row_no):
    hits=0
    for col in range(2,7):
        color=ws.cell(row_no,col).fill.fgColor
        if color.type=="theme" and color.theme==0 and color.tint < -0.20:hits+=1
    return hits>=3
def classify_payment(detail,h,i):
    text=" ".join(str(x) for x in (detail,h,i) if x is not None)
    if str(detail).strip().upper()=="R" or "リボ" in str(detail):return "リボ"
    if RATIO_RE.search(text) or END_RE.search(text):return "分割"
    return None
def extract_card_company(*values):
    text=" ".join(str(v) for v in values if v is not None)
    for alias,name in CARD_ALIASES.items():
        if alias.lower() in text.lower():return name
    return None

def parse_month_sheet(ws,month):
    income=0
    for row_no in range(1,ws.max_row+1):
        if row_no==7:continue
        v=ws.cell(row_no,2).value
        if number(v):income+=v
    payments=[];expense_total=0
    for row_no in range(1,ws.max_row+1):
        values=[ws.cell(row_no,col).value for col in range(1,min(ws.max_column,11)+1)]
        detail=ws.cell(row_no,5).value;amount=ws.cell(row_no,7).value;h=ws.cell(row_no,8).value;i=ws.cell(row_no,9).value
        if detail is None or not number(amount) or amount==0:continue
        expense_total+=amount
        category=classify_payment(detail,h,i)
        if not category:continue
        ratio=parse_ratio(h,i)
        end=END_RE.search(str(h or "")) or END_RE.search(str(i or ""))
        from_split_schedule=bool(ratio or end)
        payments.append({"name":str(detail),"type":category,"amount":amount,"progress":list(ratio) if ratio else None,"row":row_no,"fromSplitSchedule":from_split_schedule,"cardCompany":extract_card_company(*values)})
    balance=ws["B7"].value
    return income,balance if number(balance) else None,payments,expense_total

def extract_split_items(ws):
    items=[]
    if ws is None:return items
    for row_no in range(3,ws.max_row+1):
        name=ws.cell(row_no,2).value
        if name is None or is_completed_gray_row(ws,row_no):continue
        total=ws.cell(row_no,3).value;monthly=ws.cell(row_no,4).value;end=ws.cell(row_no,5).value;remaining=ws.cell(row_no,6).value
        if not number(monthly):continue
        ratio=parse_ratio(name,total,end,remaining)
        card_company=extract_card_company(name,total,end,remaining,*[ws.cell(row_no,c).value for c in range(7,min(ws.max_column,11)+1)])
        items.append({"name":str(name),"type":"リボ" if str(name).strip().upper()=="R" else "分割","total":total if number(total) else None,"monthly":monthly,"end":str(end) if end is not None else None,"remainingAmount":remaining if number(remaining) else None,"progress":list(ratio) if ratio else None,"cardCompany":card_company,"row":row_no})
    return items

def extract_life_costs(months):
    result=[]
    for m in months:
        costs={"家賃":0,"ガス":0,"電気":0,"携帯":0,"水道":0}
        for p in m["raw_expenses"]:
            for key in costs:
                if key in p["name"]:costs[key]+=p["amount"]
        result.append({"month":m["month"],**costs})
    return result

def choose_source():
    configured=os.environ.get("HOUSEHOLD_BUDGET_SOURCE")
    if configured:return Path(configured).expanduser()
    if DEFAULT_SOURCE.exists():return DEFAULT_SOURCE
    if BOOK_XLSM.exists():return BOOK_XLSM
    return BOOK_XLSX

def write_if_changed(result):
    existing=None
    if OUT.exists():
        try:existing=json.loads(OUT.read_text(encoding="utf-8"))
        except (OSError,json.JSONDecodeError):pass
    if existing:
        old_compare=dict(existing);old_compare.pop("generatedAt",None);new_compare=dict(result);new_compare.pop("generatedAt",None)
        if old_compare==new_compare:result["generatedAt"]=existing.get("generatedAt",result["generatedAt"])
    OUT.parent.mkdir(parents=True,exist_ok=True);OUT.write_text(json.dumps(result,ensure_ascii=False,indent=2)+"\n",encoding="utf-8")

def main():
    book=choose_source()
    if not book.exists():raise SystemExit(f"家計.xlsm が見つかりません: {book}")
    wb=openpyxl.load_workbook(book,data_only=True,read_only=False,keep_vba=book.suffix.lower()==".xlsm")
    months=[];split_sheet=wb["分割分"] if "分割分" in wb.sheetnames else None
    for ws in wb.worksheets:
        m=MONTH_RE.match(ws.title)
        if not m:continue
        month=f"{m.group(1)}-{int(m.group(2)):02d}"
        income,balance,payments,expense_total=parse_month_sheet(ws,month);raw_expenses=[]
        for row_no in range(1,ws.max_row+1):
            detail=ws.cell(row_no,5).value;amount=ws.cell(row_no,7).value
            if detail is not None and number(amount) and amount!=0:raw_expenses.append({"name":str(detail),"amount":amount})
        months.append({"month":month,"sheet":ws.title,"income":income,"expenseTotal":expense_total,"balance":balance,"payments":payments,"paymentTotal":sum(p["amount"] for p in payments if not p["fromSplitSchedule"]),"raw_expenses":raw_expenses})
    months.sort(key=lambda x:x["month"]);life_cost_months=extract_life_costs(months)
    for m in months:m.pop("raw_expenses",None)
    result={"generatedAt":datetime.now().isoformat(timespec="seconds"),"source":str(book),"rules":{"incomeColumn":"B","incomeRule":"各月シートのB列に記載された数値をすべて合計。ただしB7は残高セルのため収入から除外","balanceCell":"B7","paymentDetailColumn":"E","paymentAmountColumn":"G","installmentColumns":["H","I"],"splitScheduleRule":"各月シートH/I列に回/回または終了年月が記載された行は分割分シート由来として扱い、月別支払い計画の合計へ二重加算しない","completedRows":"分割分シートのグレーアウト行は除外","scheduleSource":"各カード・サービス公式HP","sourceFile":"OneDriveのデスクトップにある家計.xlsm（環境変数で変更可能）"},"cardSchedules":OFFICIAL_SCHEDULES,"serviceSchedules":OFFICIAL_SERVICES,"months":months,"lifeCostMonths":life_cost_months,"splitItems":extract_split_items(split_sheet)}
    write_if_changed(result);print(f"Generated {OUT} from {book}; monthly sheets={len(months)}; split items={len(result['splitItems'])}")
if __name__=="__main__":main()
